import re
import xlsxwriter
import openpyxl

class Employee:
    def __init__(self, name, dob, phone, email) -> None:
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email

    @staticmethod
    def get_valid_input(input_prompt, regex, error_message):
        user_input = input(input_prompt)

        while not re.fullmatch(regex, user_input):
            print(error_message)
            user_input = input(input_prompt)

        return user_input

    @staticmethod
    def get_name():
        name = Employee.get_valid_input(
            "Enter employee name: ",
            "^[A-Za-z\\s]+$",
            "Name can have only alphabets and spaces."
        )

        return name
    
    @staticmethod
    def get_dob():
        dob = Employee.get_valid_input(
            "Enter employee DOB: ",
            "^(3[01]|[12][0-9]|0[1-9])/(1[0-2]|0[1-9])/[0-9]{4}$",
            "Date of birth should be in DD/MM/YYYY format."
        )

        return dob

    @staticmethod
    def get_phone_number():
        phone_number = Employee.get_valid_input(
            "Enter employee phone number(with country code): ",
            "^[+]{1}(?:[0-9\-\\(\\)\\/.]\s?){6,15}[0-9]{1}$",
            "Enter a valid phone number."
        )

        return phone_number
    
    @staticmethod
    def get_email():
        email = Employee.get_valid_input(
            "Enter employee email address: ",
            "^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}$",
            "Enter a valid email address."
        )

        return email
    
    @staticmethod
    def get_employee_details():
        employee = Employee(
            Employee.get_name(),
            Employee.get_dob(),
            Employee.get_phone_number(),
            Employee.get_email()
        )
        
        return employee
    
class Application:
    workbook = ""
    worksheet = ""

    def __init__(self) -> None:
        Application.workbook = xlsxwriter.Workbook('employees.xlsx')
        Application.worksheet = self.workbook.add_worksheet("Employee Details")

    @staticmethod
    def get_rows_count(filename):
        workbook = openpyxl.load_workbook(filename, enumerate)
        sheet = workbook["Employee Details"]

        return sheet.max_row


    @staticmethod
    def save_employee_to_excel(employee):
        rows_count = Application.get_rows_count("employees.xlsx")
        Application.worksheet.write(rows_count, 0, employee.name)
        Application.worksheet.write(rows_count, 1, employee.dob)
        Application.worksheet.write(rows_count, 2, employee.phone)
        Application.worksheet.write(rows_count, 3, employee.email)

    @staticmethod
    def save_employee_to_textfile(employee):
        f = open('employees.txt', 'a')
        f.write(f'Name: {employee.name}\nDate of birth: {employee.dob}\nPhone number: {employee.phone}\nEmail: {employee.email}\n')
        f.close()

    def start(self):
        doContinue = 'Y'

        while doContinue == 'Y':
            employee = Employee.get_employee_details()

            Application.save_employee_to_textfile(employee)
            print("Employee added to records.")

            doContinue = input("And another employee? (Y/N)")

application = Application()
application.start()